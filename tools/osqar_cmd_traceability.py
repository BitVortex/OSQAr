"""`osqar traceability` subcommand."""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
from pathlib import Path

from tools.traceability_check import _collect_trace_links, _load_needs
from tools.traceability_check import cli as traceability_cli
from tools.typed_traceability import (
    API_AUDIT_SCHEMA,
    project_api_requirement_paths,
    project_api_requirements,
    validate_typed_traceability,
)

API_REQUIREMENT_FIELDS = [
    "API_ID",
    "API_Title",
    "Requirement_IDs",
    "Requirement_Titles",
    "Allocation_Status",
    "Profile",
    "Schema",
]


def _paths_alias(left: Path, right: Path) -> bool:
    """Return whether two paths name the same lexical or existing filesystem object."""
    if left == right:
        return True
    try:
        return left.exists() and right.exists() and left.samefile(right)
    except OSError:
        return False


def _run_typed_traceability(args: argparse.Namespace) -> int:
    needs_path = Path(str(args.needs_json)).expanduser().resolve()
    profile = str(getattr(args, "profile", "basic"))
    artifact_arg = getattr(args, "api_requirements_output", None)
    artifact = Path(artifact_arg).expanduser().resolve() if artifact_arg else None
    audit = artifact.with_suffix(".audit.json") if artifact is not None else None
    report_arg = getattr(args, "json_report", None)
    report_path = Path(report_arg).expanduser().resolve() if report_arg else None
    evidence_arg = getattr(args, "evidence_project", None)
    evidence_project = Path(evidence_arg).expanduser().resolve() if evidence_arg else None
    validation_context = {
        "evidence_project": evidence_project,
        "expected_source_revision": getattr(args, "source_revision", None),
        "expected_configuration_sha256": getattr(args, "configuration_sha256", None),
    }
    api_prefixes = tuple(getattr(args, "api_prefix", None) or ("API_", "IMPL_"))

    outputs = tuple(path for path in (artifact, audit, report_path) if path is not None)
    protected_inputs = tuple(
        path for path in (needs_path, evidence_project) if path is not None
    )
    for output in outputs:
        for protected in protected_inputs:
            if _paths_alias(output, protected):
                print(
                    f"ERROR: traceability output {output} aliases input {protected}",
                    file=sys.stderr,
                )
                return 2
    for index, output in enumerate(outputs):
        for other in outputs[index + 1:]:
            if _paths_alias(output, other):
                print(
                    f"ERROR: traceability outputs alias each other: {output} and {other}",
                    file=sys.stderr,
                )
                return 2

    # A failed invocation must not leave stale files that appear to belong to it.
    try:
        for output in outputs:
            output.unlink(missing_ok=True)
    except OSError as exc:
        print(f"ERROR: failed to invalidate stale traceability output: {exc}", file=sys.stderr)
        return 2

    if artifact is not None and profile != "qualification":
        print(
            "ERROR: --api-requirements-output requires --profile qualification",
            file=sys.stderr,
        )
        return 2

    try:
        needs = _load_needs(needs_path)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: Failed to read {needs_path}: {exc}", file=sys.stderr)
        return 2

    if profile == "qualification":
        report = validate_typed_traceability(needs, profile=profile, **validation_context)
        if report_path is not None:
            try:
                report_path.parent.mkdir(parents=True, exist_ok=True)
                report_path.write_text(
                    json.dumps(report.as_dict(), indent=2, sort_keys=True) + "\n",
                    encoding="utf-8",
                )
            except OSError as exc:
                print(f"ERROR: failed to write typed traceability report: {exc}", file=sys.stderr)
                return 2
        if report.status != "PASS":
            print(
                f"Typed traceability: FAIL ({len(report.violations)} violation(s))",
                file=sys.stderr,
            )
            for violation in report.violations:
                print(f"  - {violation}", file=sys.stderr)
            return 1
        print(f"Typed traceability: PASS ({len(report.executed_rules)} rules)")

    if artifact is not None and audit is not None:
        try:
            rows = project_api_requirements(
                needs,
                profile=profile,
                api_prefixes=api_prefixes,
                **validation_context,
            )
            paths = project_api_requirement_paths(
                needs,
                profile=profile,
                api_prefixes=api_prefixes,
                **validation_context,
            )
            csv_buffer = io.StringIO(newline="")
            writer = csv.DictWriter(csv_buffer, fieldnames=API_REQUIREMENT_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
            audit_text = json.dumps(
                {
                    "schema": API_AUDIT_SCHEMA,
                    "profile": profile,
                    "paths": paths,
                    "limitation": (
                        "The CSV collapses architecture for presentation; this audit "
                        "file preserves authored intermediate paths."
                    ),
                },
                indent=2,
                sort_keys=True,
            ) + "\n"
            artifact.parent.mkdir(parents=True, exist_ok=True)
            audit.parent.mkdir(parents=True, exist_ok=True)
            artifact.write_text(csv_buffer.getvalue(), encoding="utf-8")
            audit.write_text(audit_text, encoding="utf-8")
        except (OSError, ValueError) as exc:
            for output in outputs:
                try:
                    output.unlink(missing_ok=True)
                except OSError:
                    pass
            print(f"ERROR: failed to write API allocation artifacts: {exc}", file=sys.stderr)
            return 2
        print(
            f"API-to-requirement allocation artifact written: {artifact} "
            f"({len(rows)} APIs); audit paths: {audit}"
        )
    return 0


def _need_title(need: dict) -> str:
    """Extract a human-readable title from a need."""
    title = str(need.get("title", need.get("content", ""))).strip()
    if not title:
        title = str(need.get("id", ""))
    # Clean: collapse whitespace, remove embedded newlines
    title = " ".join(title.split())
    return title


def _normalize_links(links: object) -> str:
    """Convert links field to semicolon-delimited string."""
    if links is None:
        return ""
    if isinstance(links, str):
        return links
    if isinstance(links, (list, tuple)):
        return ";".join(str(l) for l in links if l)
    return str(links)


def _export_csv(
    needs_json_path: Path,
    output_path: Path,
    *,
    req_prefixes: tuple[str, ...] = ("REQ_",),
    arch_prefixes: tuple[str, ...] = ("ARCH_",),
    test_prefixes: tuple[str, ...] = ("TEST_", "VER_"),
    code_prefixes: tuple[str, ...] = ("CODE_", "IMPL_"),
    lm_prefixes: tuple[str, ...] = ("LM_",),
) -> int:
    """Export a traceability matrix as CSV from needs.json."""
    try:
        needs = _load_needs(needs_json_path)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: Failed to read {needs_json_path}: {exc}")
        return 2

    if not needs:
        print("WARNING: no needs found in needs.json")
        return 1

    # Build lookup
    needs_by_id: dict[str, dict] = {str(n.get("id", "")): n for n in needs if n.get("id")}
    all_ids = set(needs_by_id)

    # Identify requirements (rows) and all linked types (columns)
    req_rows: list[dict] = []
    for nid, need in sorted(needs_by_id.items()):
        if nid.startswith(req_prefixes):
            req_rows.append(need)

    if not req_rows:
        print("WARNING: no requirements found (check --req-prefix)")
        return 1

    # CSV columns
    fieldnames = [
        "REQ_ID",
        "REQ_Title",
        "Status",
        "Tags",
        "ARCH_Linked",
        "VER_Linked",
        "IMPL_Linked",
        "LM_Linked",
        "Other_Linked",
        "Total_Links",
    ]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for req in req_rows:
        nid = str(req.get("id", ""))
        linked = _collect_trace_links(req)

        # Classify links by target prefix
        arch_links: list[str] = []
        ver_links: list[str] = []
        impl_links: list[str] = []
        lm_links: list[str] = []
        other_links: list[str] = []

        for target in sorted(linked):
            if not target:
                continue
            if target in all_ids and target not in needs_by_id:
                continue  # dead link — skip
            if target.startswith(arch_prefixes):
                arch_links.append(target)
            elif target.startswith(test_prefixes):
                ver_links.append(target)
            elif target.startswith(code_prefixes):
                impl_links.append(target)
            elif target.startswith(lm_prefixes):
                lm_links.append(target)
            else:
                other_links.append(target)

        tags = req.get("tags")
        if isinstance(tags, list):
            tags_str = ";".join(str(t) for t in tags if t)
        elif isinstance(tags, str):
            tags_str = tags
        else:
            tags_str = ""

        row = {
            "REQ_ID": nid,
            "REQ_Title": _need_title(req),
            "Status": str(req.get("status", "")),
            "Tags": tags_str,
            "ARCH_Linked": ";".join(arch_links),
            "VER_Linked": ";".join(ver_links),
            "IMPL_Linked": ";".join(impl_links),
            "LM_Linked": ";".join(lm_links),
            "Other_Linked": ";".join(other_links),
            "Total_Links": str(len(linked)),
        }
        writer.writerow(row)

    csv_text = buf.getvalue()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(csv_text, encoding="utf-8")

    print(f"Traceability matrix written: {output_path}")
    print(f"  {len(req_rows)} requirements, CSV columns: {', '.join(fieldnames)}")
    return 0


def _export_xlsx(
    needs_json_path: Path,
    output_path: Path,
    *,
    req_prefixes: tuple[str, ...] = ("REQ_",),
    arch_prefixes: tuple[str, ...] = ("ARCH_",),
    test_prefixes: tuple[str, ...] = ("TEST_", "VER_"),
    code_prefixes: tuple[str, ...] = ("CODE_", "IMPL_"),
    lm_prefixes: tuple[str, ...] = ("LM_",),
) -> int:
    """Export a traceability matrix as .xlsx from needs.json."""
    try:
        import openpyxl
    except ImportError:
        print(
            "ERROR: openpyxl is required for XLSX export. Install it with:\n"
            "  pip install openpyxl",
            file=__import__("sys").stderr,
        )
        return 2

    try:
        needs = _load_needs(needs_json_path)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: Failed to read {needs_json_path}: {exc}")
        return 2

    if not needs:
        print("WARNING: no needs found in needs.json")
        return 1

    needs_by_id: dict[str, dict] = {str(n.get("id", "")): n for n in needs if n.get("id")}
    req_rows: list[dict] = []
    for nid, need in sorted(needs_by_id.items()):
        if nid.startswith(req_prefixes):
            req_rows.append(need)

    if not req_rows:
        print("WARNING: no requirements found (check --req-prefix)")
        return 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traceability Matrix"

    # Header
    headers = [
        "REQ_ID", "REQ_Title", "Status", "Tags",
        "ARCH_Linked", "VER_Linked", "IMPL_Linked", "LM_Linked",
        "Other_Linked", "Total_Links",
    ]
    header_font = openpyxl.styles.Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Data rows
    for row_idx, req in enumerate(req_rows, 2):
        nid = str(req.get("id", ""))
        linked = _collect_trace_links(req)

        arch_links: list[str] = []
        ver_links: list[str] = []
        impl_links: list[str] = []
        lm_links: list[str] = []
        other_links: list[str] = []

        for target in sorted(linked):
            if not target:
                continue
            if target.startswith(arch_prefixes):
                arch_links.append(target)
            elif target.startswith(test_prefixes):
                ver_links.append(target)
            elif target.startswith(code_prefixes):
                impl_links.append(target)
            elif target.startswith(lm_prefixes):
                lm_links.append(target)
            else:
                other_links.append(target)

        tags = req.get("tags")
        if isinstance(tags, list):
            tags_str = ";".join(str(t) for t in tags if t)
        elif isinstance(tags, str):
            tags_str = tags
        else:
            tags_str = ""

        row_data = [
            nid,
            _need_title(req),
            str(req.get("status", "")),
            tags_str,
            ";".join(arch_links),
            ";".join(ver_links),
            ";".join(impl_links),
            ";".join(lm_links),
            ";".join(other_links),
            len(linked),
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val)

    # Auto-fit column widths (approximate)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    print(f"Traceability matrix (XLSX) written: {output_path}")
    print(f"  {len(req_rows)} requirements, columns: {', '.join(headers)}")
    return 0


def cmd_traceability(args: argparse.Namespace) -> int:
    profile = str(getattr(args, "profile", "basic"))
    if profile == "qualification" or getattr(args, "api_requirements_output", None):
        return _run_typed_traceability(args)

    # XLSX export mode
    fmt = getattr(args, "format", None)
    if fmt == "xlsx":
        output = getattr(args, "format_output", None)
        if not output:
            print(
                "ERROR: --format-output is required with --format xlsx",
                __import__("sys").stderr,
            )
            return 2
        output_path = Path(output).expanduser().resolve()
        return _export_xlsx(
            Path(str(args.needs_json)).expanduser().resolve(),
            output_path,
            req_prefixes=tuple(getattr(args, "req_prefix", []) or ["REQ_"]),
            arch_prefixes=tuple(getattr(args, "arch_prefix", []) or ["ARCH_"]),
            test_prefixes=tuple(getattr(args, "test_prefix", []) or ["TEST_", "VER_"]),
            code_prefixes=tuple(getattr(args, "code_prefix", []) or ["CODE_", "IMPL_"]),
            lm_prefixes=tuple(getattr(args, "lm_prefix", []) or ["LM_"]),
        )

    # CSV export mode
    fmt = getattr(args, "format", None)
    if fmt == "csv":
        output = getattr(args, "format_output", None)
        if not output:
            print("ERROR: --format-output is required with --format csv", __import__("sys").stderr)
            return 2
        output_path = Path(output).expanduser().resolve()
        return _export_csv(
            Path(str(args.needs_json)).expanduser().resolve(),
            output_path,
            req_prefixes=tuple(getattr(args, "req_prefix", []) or ["REQ_"]),
            arch_prefixes=tuple(getattr(args, "arch_prefix", []) or ["ARCH_"]),
            test_prefixes=tuple(getattr(args, "test_prefix", []) or ["TEST_", "VER_"]),
            code_prefixes=tuple(getattr(args, "code_prefix", []) or ["CODE_", "IMPL_"]),
            lm_prefixes=tuple(getattr(args, "lm_prefix", []) or ["LM_"]),
        )

    # Standard traceability check
    argv: list[str] = [str(args.needs_json)]
    if getattr(args, "json_report", None):
        argv += ["--json-report", str(args.json_report)]
    if getattr(args, "project_config", None):
        argv += ["--project-config", str(args.project_config)]

    for prefix_attr in ("req_prefix", "arch_prefix", "test_prefix", "code_prefix"):
        values = getattr(args, prefix_attr, None) or []
        flag = "--" + prefix_attr.replace("_", "-")
        for v in values:
            argv.extend([flag, str(v)])

    if getattr(args, "enforce_req_has_test", False):
        argv += ["--enforce-req-has-test"]
    if getattr(args, "enforce_arch_traces_req", False):
        argv += ["--enforce-arch-traces-req"]
    if getattr(args, "enforce_test_traces_req", False):
        argv += ["--enforce-test-traces-req"]

    return int(traceability_cli(argv))


def register(sub: argparse._SubParsersAction) -> None:
    p_tr = sub.add_parser(
        "traceability", help="Run traceability checks or export traceability matrix"
    )
    p_tr.add_argument("needs_json", type=Path, help="Path to needs.json or needs.yaml")
    p_tr.add_argument(
        "--json-report", type=Path, default=None, help="Write JSON report to this path"
    )
    p_tr.add_argument(
        "--project-config",
        type=Path,
        default=None,
        help="Project configuration declaring standards catalogs for STDCLAIM_* needs",
    )
    p_tr.add_argument(
        "--profile",
        choices=["basic", "qualification"],
        default="basic",
        help="Validation behavior profile (default: basic compatibility mode)",
    )
    p_tr.add_argument(
        "--api-requirements-output",
        type=Path,
        default=None,
        help=(
            "Optionally write a CSV API-to-requirement allocation view. The "
            "projection traverses requirement -> architecture -> API backwards "
            "but omits architecture from the displayed artifact."
        ),
    )
    p_tr.add_argument(
        "--api-prefix",
        action="append",
        default=None,
        help=(
            "Implementation ID prefix treated as an API in the qualification "
            "projection (repeatable; default: API_, IMPL_)"
        ),
    )
    p_tr.add_argument(
        "--evidence-project",
        type=Path,
        default=None,
        help="Qualification evidence project accepted by the framework validator",
    )
    p_tr.add_argument(
        "--source-revision",
        default=None,
        help="Independent trusted source revision for qualification evidence",
    )
    p_tr.add_argument(
        "--configuration-sha256",
        default=None,
        help="Independent trusted configuration SHA-256 for qualification evidence",
    )
    p_tr.add_argument(
        "--format",
        choices=["csv", "xlsx"],
        default="check",
        help="Output mode: 'check' (default) for violations, 'csv' for traceability matrix export, 'xlsx' for Excel spreadsheet export",
    )
    p_tr.add_argument(
        "--format-output",
        type=Path,
        default=None,
        help="File to write when using --format csv (required for CSV)",
    )
    p_tr.add_argument(
        "--enforce-req-has-test",
        action="store_true",
        help="Also enforce REQ_* → TEST_* coverage",
    )
    p_tr.add_argument(
        "--enforce-arch-traces-req",
        action="store_true",
        help="Also enforce ARCH_* → REQ_* coverage",
    )
    p_tr.add_argument(
        "--enforce-test-traces-req",
        action="store_true",
        help="Also enforce TEST_* → REQ_* coverage",
    )
    p_tr.add_argument(
        "--req-prefix",
        action="append",
        default=[],
        help="Requirement ID prefix (repeatable; default: REQ_)",
    )
    p_tr.add_argument(
        "--arch-prefix",
        action="append",
        default=[],
        help="Architecture ID prefix (repeatable; default: ARCH_)",
    )
    p_tr.add_argument(
        "--test-prefix",
        action="append",
        default=[],
        help="Test ID prefix (repeatable; default: TEST_)",
    )
    p_tr.add_argument(
        "--code-prefix",
        action="append",
        default=[],
        help="Implementation/code ID prefix (repeatable; default: CODE_, IMPL_)",
    )
    p_tr.add_argument(
        "--lm-prefix",
        action="append",
        default=[],
        help="Lifecycle management ID prefix (repeatable; default: LM_)",
    )
    p_tr.set_defaults(func=cmd_traceability)
